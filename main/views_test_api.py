from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet
import qrcode
from qrcode.constants import ERROR_CORRECT_H
import io
from PIL import Image, ImageDraw, ImageFont
from django.shortcuts import render
from .models import StudentTest, StudentAnswer, PdfVerification
from .models import StudentTest, StudentAnswer, PdfVerification, Group, Kafedra, Bulim
# Fan bo'yicha PDF natija yuklash
from django.utils.encoding import smart_str
def export_subject_results_pdf(request, subject_name):
    from datetime import datetime
    # Asosiy queryset
    tests_qs = StudentTest.objects.filter(test__subject__name=subject_name, completed=True).select_related('student', 'test', 'test__group')

    # --- Query parametrlardan filtrlar ---
    group_name = request.GET.get('group') or ''
    semester_number = request.GET.get('semester') or ''
    attempt_count_param = request.GET.get('attempt_count') or ''  # exact
    attempt_gte_param = request.GET.get('attempt_gte') or ''       # >=
    attempt_min_param = request.GET.get('attempt_min') or ''       # range min
    attempt_max_param = request.GET.get('attempt_max') or ''       # range max
    attempt_nth_param = request.GET.get('attempt_nth') or ''       # specific nth attempt

    if group_name:
        tests_qs = tests_qs.filter(test__group__name=group_name)
    if semester_number:
        # Aniqroq semestr filtri: GroupSubject orqali group id larini topamiz
        from main.models import GroupSubject
        if group_name:
            group_subject_ids = GroupSubject.objects.filter(
                group__name=group_name,
                subject__name=subject_name,
                semester__number=semester_number
            ).values_list('group_id', flat=True)
            tests_qs = tests_qs.filter(test__group_id__in=group_subject_ids)
        else:
            group_ids = GroupSubject.objects.filter(
                subject__name=subject_name,
                semester__number=semester_number
            ).values_list('group_id', flat=True)
            tests_qs = tests_qs.filter(test__group_id__in=group_ids)
    tests_qs = tests_qs.distinct()

    # Attempt (urinish) filtrlash rejimlari:
    # 1) attempt_nth_param: har bir talabadan faqat n-chi urinish (agar mavjud bo'lsa)
    # 2) attempt_count_param: aynan n marta topshirgan talabalar (oxirgi urinish)
    # 3) attempt_gte_param: kamida n marta topshirganlar (oxirgi urinish)
    # 4) attempt_min_param / attempt_max_param: oraliq (oxirgi urinish)
    from collections import defaultdict
    bucket = defaultdict(list)
    ordered_qs = tests_qs.order_by('student_id', 'start_time')
    for st in ordered_qs:
        bucket[st.student_id].append(st)

    tests = []
    attempt_count_val = None  # exact
    attempt_gte_val = None
    attempt_min_val = None
    attempt_max_val = None
    attempt_nth_val = None

    # Parse ints safely
    def to_int(val):
        try:
            return int(val)
        except (TypeError, ValueError):
            return None

    attempt_nth_val = to_int(attempt_nth_param)
    if attempt_nth_val and attempt_nth_val > 0:
        # Faqat n-chi urinish
        for arr in bucket.values():
            if len(arr) >= attempt_nth_val:
                tests.append(arr[attempt_nth_val-1])
    else:
        attempt_count_val = to_int(attempt_count_param)
        attempt_gte_val = to_int(attempt_gte_param)
        attempt_min_val = to_int(attempt_min_param)
        attempt_max_val = to_int(attempt_max_param)
        for arr in bucket.values():
            total_attempts = len(arr)
            ok = True
            if attempt_count_val and total_attempts != attempt_count_val:
                ok = False
            if ok and attempt_gte_val and total_attempts < attempt_gte_val:
                ok = False
            if ok and attempt_min_val and total_attempts < attempt_min_val:
                ok = False
            if ok and attempt_max_val and total_attempts > attempt_max_val:
                ok = False
            if ok:
                tests.append(arr[-1])  # oxirgi urinish
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{smart_str(subject_name)}_test_natijalari.pdf"'

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.styles import ParagraphStyle

    # Custom styles
    title_style = ParagraphStyle('title', parent=styles['Title'], alignment=TA_CENTER, fontSize=16, spaceAfter=8)
    subtitle_style = ParagraphStyle('subtitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=13, spaceAfter=8)
    normal_style = ParagraphStyle('normal', parent=styles['Normal'], fontSize=11, spaceAfter=4)
    right_style = ParagraphStyle('right', parent=styles['Normal'], alignment=TA_RIGHT, fontSize=11)
    left_style = ParagraphStyle('left', parent=styles['Normal'], alignment=TA_LEFT, fontSize=11)

    # Header (title, date right, subtitle)
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Kattaqurg'on Davlat Pedagogika instituti", ParagraphStyle('header', parent=styles['Normal'], alignment=TA_CENTER, fontSize=14, spaceAfter=0, leading=16)))
    # Eng erta real boshlanish vaqtini aniqlash (filtrlangan ro'yxat ichida)
    from django.utils.timezone import localtime
    if tests:
        earliest = min([t.start_time for t in tests if t.start_time]) if any(t.start_time for t in tests) else None
    else:
        earliest = None
    if earliest:
        test_date_display = localtime(earliest).strftime('%d.%m.%Y')
    else:
        test_date_display = datetime.now().strftime('%d.%m.%Y')
    elements.append(Paragraph("Yakuniy nazorat test sinovlari natijalari", ParagraphStyle('subtitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=12, spaceAfter=2, leading=14)))
    elements.append(Spacer(1, 32))
    elements.append(Paragraph(f"Fanning nomi: {subject_name}", ParagraphStyle('subj', parent=styles['Normal'], fontSize=11, alignment=TA_LEFT, spaceAfter=2)))
    elements.append(Paragraph(f"Test o'tkazilgan sana: {test_date_display}", ParagraphStyle('dateHead', parent=styles['Normal'], fontSize=9, alignment=TA_LEFT, textColor=colors.black)))
    elements.append(Spacer(1, 12))

    # Filter summary (foydalanuvchi tanlagan parametrlar)
    filter_bits = []
    if group_name:
        filter_bits.append(f"Guruh: {group_name}")
    if semester_number:
        filter_bits.append(f"Semestr: {semester_number}")
    if attempt_nth_val:
        filter_bits.append(f"{attempt_nth_val}-urinish natijalari")
    else:
        if attempt_count_val:
            filter_bits.append(f"Aynan {attempt_count_val} marta (oxirgi urinish)")
        if attempt_gte_val:
            filter_bits.append(f">= {attempt_gte_val} marta (oxirgi urinish)")
        if attempt_min_val or attempt_max_val:
            span_min = attempt_min_val if attempt_min_val else 1
            span_max = attempt_max_val if attempt_max_val else '∞'
            filter_bits.append(f"Oraliq: {span_min}–{span_max} marta (oxirgi urinish)")
  
    # Table header and data
    is_super = request.user.is_authenticated and request.user.is_superuser
    if is_super:
        data = [[
            Paragraph('<b>№</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>F.I.O</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>Guruh</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>Savollar soni</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>To\'g\'ri javoblar</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>Asl foiz</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>Yakuniy foiz</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>Final ball</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>O\'tgan?</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>Status</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
        ]]
    else:
        data = [[
            Paragraph('<b>№</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>F.I.O</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>Guruh</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>Savollar soni</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>To\'g\'ri javoblar</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
            Paragraph('<b>Foizi</b>', ParagraphStyle('th', alignment=TA_CENTER, fontSize=10)),
        ]]
    for idx, stest in enumerate(tests, 1):
        fio = f"{stest.student.last_name.upper()} {stest.student.first_name.upper()} {getattr(stest.student, 'middle_name', '')}".strip()
        group = stest.test.group.name if stest.test.group else "-"
        question_ids = stest.question_ids if stest.question_ids else []
        if question_ids:
            answers = StudentAnswer.objects.filter(student_test=stest, question_id__in=question_ids)
        else:
            answers = StudentAnswer.objects.filter(student_test=stest)
        total = answers.count()
        correct = answers.filter(is_correct=True).count()
        # Original foiz (savollarning to'g'ri javobidan kelib chiqqan holda)
        original_percent = (correct/total)*100 if total else 0
        # Yakuniy foiz (override bo'lsa final_score / test.total_score)
        if hasattr(stest, 'final_score'):
            try:
                final_percent = (stest.final_score / stest.test.total_score) * 100 if stest.test.total_score else original_percent
            except Exception:
                final_percent = original_percent
        else:
            final_percent = original_percent
        original_percent_str = f"{original_percent:.1f}".replace('.', ',') + "%"
        final_percent_str = f"{final_percent:.1f}".replace('.', ',') + "%"
        if is_super:
            status_text = "Override" if (getattr(stest, 'overridden_score', None) is not None or getattr(stest, 'pass_override', False)) else "Normal"
            final_ball_str = f"{getattr(stest,'final_score', stest.total_score):.1f}".replace('.', ',')
            final_passed = getattr(stest, 'final_passed', False)
            passed_label = "Ha" if final_passed else "Yo'q"
            data.append([
                idx,
                Paragraph(fio, ParagraphStyle('td', alignment=TA_LEFT, fontSize=10)),
                group,
                total,
                correct,
                original_percent_str,
                final_percent_str,
                final_ball_str,
                passed_label,
                status_text
            ])
        else:
            data.append([
                idx,
                Paragraph(fio, ParagraphStyle('td', alignment=TA_LEFT, fontSize=10)),
                group,
                total,
                correct,
                final_percent_str
            ])

    # Table column widths (ixcham va bir xil)
    if is_super:
        table = Table(data, colWidths=[8*mm, 45*mm, 20*mm, 18*mm, 18*mm, 18*mm, 18*mm, 18*mm, 15*mm, 18*mm])
    else:
        table = Table(data, colWidths=[13*mm, 55*mm, 28*mm, 28*mm, 38*mm, 22*mm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),  # №
        ('ALIGN', (1,0), (1,-1), 'LEFT'),    # F.I.O
        ('ALIGN', (2,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,0), (-1,0), 1, colors.black),
        ('LINEABOVE', (0,0), (-1,0), 1, colors.black),
        ('LINEBEFORE', (0,0), (0,-1), 1, colors.black),
        ('LINEAFTER', (-1,0), (-1,-1), 1, colors.black),
        ('LINEBELOW', (0,-1), (-1,-1), 1, colors.black),
        ('LINEABOVE', (0,1), (-1,1), 0.5, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    # Imzo blokini jadval tugagan joydan keyin (foydalanuvchi so'rovi bo'yicha)
    elements.append(Spacer(1, 8))
    # Imzo satri: O'UBB:  __________(imzoga joy)  I.Madatov
    sig_mid_style = ParagraphStyle('sigmid', parent=left_style, alignment=TA_CENTER, fontSize=10, spaceAfter=0)
    sig_left_style = ParagraphStyle('sigleft', parent=left_style, fontSize=10, spaceAfter=0)
    sig_name_style = ParagraphStyle('signame', parent=left_style, fontSize=10, spaceAfter=0)
    signature_data = [[
        Paragraph("O'UBB:", sig_left_style),
        Paragraph("____________", sig_mid_style),
        Paragraph("I.Madatov", sig_name_style)
    ]]
    signature_table = Table(signature_data, colWidths=[25*mm, 70*mm, 35*mm])
    signature_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
        ('ALIGN', (2,0), (2,0), 'LEFT'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        # No borders
        ('BOX', (0,0), (-1,-1), 0, colors.white),
        ('INNERGRID', (0,0), (-1,-1), 0, colors.white),
    ]))
    elements.append(signature_table)
    elements.append(Spacer(1, 16))

    # QR kod va imzo footerga (past) joylashtiriladi
    import hashlib, time
    record_count = len(tests)
    payload_raw = f"SUBJECT={subject_name};COUNT={record_count};TS={int(time.time())}"
    sig_hash = hashlib.sha256(payload_raw.encode()).hexdigest()[:32]
    verification_obj, created = PdfVerification.objects.get_or_create(
        hash_code=sig_hash,
        defaults={
            'subject_name': subject_name,
            'record_count': record_count,
            'payload': payload_raw,
            'generated_by': request.user if request.user.is_authenticated else None
        }
    )
    qr_text = request.build_absolute_uri(f"/api/test-api/verify-qr/{verification_obj.hash_code}/")
    # Yangi: yuqori xatolik tuzatish (H) va avtomatik versiya tanlash, biroz katta modul
    qr = qrcode.QRCode(version=None, error_correction=ERROR_CORRECT_H, box_size=4, border=2)
    qr.add_data(qr_text)
    qr.make(fit=True)
    # Rangli modul va fon (yuqori kontrast saqlanadi)
    module_color = (19, 46, 120)      # chuqur ko'k
    background_color = (255, 255, 255)
    img = qr.make_image(fill_color=module_color, back_color=background_color).convert('RGBA')

    # --- Markaziy 'RTTM' overlay (gradient bilan) ---
    draw = ImageDraw.Draw(img)
    W, H = img.size
    overlay_ratio = 0.18  # modul maydonining kichik qismi
    base_size = int(W * overlay_ratio)
    try:
        font = ImageFont.truetype("arial.ttf", base_size)
    except Exception:
        font = ImageFont.load_default()
    text = "RTTM"
    tb = draw.textbbox((0,0), text, font=font)
    tw, th = tb[2]-tb[0], tb[3]-tb[1]
    pad_x = 8
    pad_y = 6
    box_x0 = (W - tw)//2 - pad_x
    box_y0 = (H - th)//2 - pad_y
    box_x1 = box_x0 + tw + pad_x*2
    box_y1 = box_y0 + th + pad_y*2
    # Gradient tayyorlash
    grad_w = int(box_x1 - box_x0)
    grad_h = int(box_y1 - box_y0)
    from math import sqrt
    gradient = Image.new('RGBA', (grad_w, grad_h))
    gdraw = ImageDraw.Draw(gradient)
    # Gradient ranglari (chapdan o'ngga ko'k -> binafsha)
    start_col = (37, 99, 235)
    end_col = (147, 51, 234)
    for x in range(grad_w):
        t = x / max(1, grad_w-1)
        r = int(start_col[0] + (end_col[0]-start_col[0])*t)
        g = int(start_col[1] + (end_col[1]-start_col[1])*t)
        b = int(start_col[2] + (end_col[2]-start_col[2])*t)
        gdraw.line([(x,0),(x,grad_h)], fill=(r,g,b,235))
    # Yumaloq mask
    mask = Image.new('L', (grad_w, grad_h), 0)
    mask_draw = ImageDraw.Draw(mask)
    try:
        mask_draw.rounded_rectangle([0,0,grad_w,grad_h], radius=10, fill=255)
    except Exception:
        mask_draw.rectangle([0,0,grad_w,grad_h], fill=255)
    # Soyali effekt (shadow) orqa fon uchun
    shadow = Image.new('RGBA', (grad_w+6, grad_h+6), (0,0,0,0))
    sdraw = ImageDraw.Draw(shadow)
    sdraw.ellipse([3,3,grad_w+3,grad_h+3], fill=(0,0,0,60))
    img.alpha_composite(shadow, (int(box_x0-3), int(box_y0-3)))
    img.paste(gradient, (int(box_x0), int(box_y0)), mask)
    # Matn (oq, ozgina soyali)
    text_x = (W - tw)//2
    text_y = (H - th)//2
    # Soyasi (1px)
    draw.text((text_x+1, text_y+1), text, font=font, fill=(0,0,0,90))
    draw.text((text_x, text_y), text, font=font, fill=(255,255,255,240))
    qr_buffer = io.BytesIO()
    img.save(qr_buffer, format='PNG')

    def footer(c, doc):
        # Faqat QR kodni pastki o'ng burchakka joylashtirish
        from reportlab.lib.pagesizes import A4 as _A4
        size = 25*mm
        qr_buffer.seek(0)
        c.drawImage(ImageReader(qr_buffer), _A4[0]-doc.rightMargin-size, doc.bottomMargin, size, size, preserveAspectRatio=True, mask='auto')

    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


def verify_qr(request, hash_code):
    """QR kod orqali kelgan hashni tekshirish.
    Topilsa: ma'lumot + 'Haqiqiy'. Topilmasa: 'Noto'g'ri' xabari.
    """
    obj = PdfVerification.objects.filter(hash_code=hash_code).first()
    if not obj:
        return render(request, 'test_api/verify_qr.html', {
            'status': 'invalid',
            'hash': hash_code,
        })
    return render(request, 'test_api/verify_qr.html', {
        'status': 'valid',
        'hash': obj.hash_code,
        'subject': obj.subject_name,
        'count': obj.record_count,
        'created_at': obj.created_at,
        'payload': obj.payload,
    })
from django.contrib.auth.decorators import login_required
@login_required
def testapi_logout(request):
    logout(request)
    return redirect('testapi_login')
# Xodimlar natijalarini bo‘lim bo‘yicha eksport
def export_employees_by_bulim_excel(request, bulim_id):
    if not request.user.is_authenticated:
        return redirect('/api/login/')
    from main.models import Bulim
    try:
        bulim = Bulim.objects.get(id=bulim_id)
    except Bulim.DoesNotExist:
        return HttpResponse("Bo'lim topilmadi", status=404)
    employee_users = User.objects.filter(role='employee', bulim=bulim)
    employee_tests = StudentTest.objects.filter(completed=True, student__in=employee_users).select_related('student', 'test', 'test__subject', 'test__group').order_by('student__username', '-start_time')
    wb = Workbook()
    ws = wb.active
    ws.title = f"{bulim.name} - Xodimlar"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    headers = ["Xodim F.I.Sh.", "Username", "Test", "Fan", "Test sanasi", "Savollar soni", "To'g'ri javob", "Xato javob", "Ball", "Maksimal ball", "Foiz"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    row = 2
    for stest in employee_tests:
        fio = f"{stest.student.first_name} {stest.student.last_name}"
        username = stest.student.username
        test_name = stest.test.subject.name if stest.test.subject else "-"
        subject = stest.test.subject.name if stest.test.subject else "-"
        test_date = stest.start_time.strftime("%d.%m.%Y %H:%M")
        question_ids = stest.question_ids if stest.question_ids else []
        if question_ids:
            answers = StudentAnswer.objects.filter(student_test=stest, question_id__in=question_ids)
        else:
            answers = StudentAnswer.objects.filter(student_test=stest)
        total = answers.count()
        correct = answers.filter(is_correct=True).count()
        incorrect = total - correct
        score = sum([a.score for a in answers])
        percent = int((correct / total) * 100) if total else 0
        data = [fio, username, test_name, subject, test_date, total, correct, incorrect, score, stest.test.total_score, f"{percent}%"]
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1
    column_widths = [22, 15, 20, 18, 18, 10, 12, 12, 10, 12, 8]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{bulim.name}_xodimlar_natijalari.xlsx"'
    wb.save(response)
    return response
# Tutorlarnatijalarini kafedra bo‘yicha eksport
def export_tutors_by_kafedra_excel(request, kafedra_id):
    if not request.user.is_authenticated:
        return redirect('/api/login/')
    from main.models import Kafedra
    try:
        kafedra = Kafedra.objects.get(id=kafedra_id)
    except Kafedra.DoesNotExist:
        return HttpResponse("Kafedra topilmadi", status=404)
    tutor_users = User.objects.filter(role='tutor', kafedra=kafedra)
    tutor_tests = StudentTest.objects.filter(completed=True, student__in=tutor_users).select_related('student', 'test', 'test__subject', 'test__group').order_by('student__username', '-start_time')
    wb = Workbook()
    ws = wb.active
    ws.title = f"{kafedra.name} - Tutorlar"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    headers = ["Tutor F.I.Sh.", "Username", "Test", "Fan", "Test sanasi", "Savollar soni", "To'g'ri javob", "Xato javob", "Ball", "Maksimal ball", "Foiz"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    row = 2
    for stest in tutor_tests:
        fio = f"{stest.student.first_name} {stest.student.last_name}"
        username = stest.student.username
        test_name = stest.test.subject.name if stest.test.subject else "-"
        subject = stest.test.subject.name if stest.test.subject else "-"
        test_date = stest.start_time.strftime("%d.%m.%Y %H:%M")
        question_ids = stest.question_ids if stest.question_ids else []
        if question_ids:
            answers = StudentAnswer.objects.filter(student_test=stest, question_id__in=question_ids)
        else:
            answers = StudentAnswer.objects.filter(student_test=stest)
        total = answers.count()
        correct = answers.filter(is_correct=True).count()
        incorrect = total - correct
        score = sum([a.score for a in answers])
        percent = int((correct / total) * 100) if total else 0
        data = [fio, username, test_name, subject, test_date, total, correct, incorrect, score, stest.test.total_score, f"{percent}%"]
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1
    column_widths = [22, 15, 20, 18, 18, 10, 12, 12, 10, 12, 8]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{kafedra.name}_tutorlar_natijalari.xlsx"'
    wb.save(response)
    return response
# Talabalar natijalarini guruh bo‘yicha eksport
def export_students_by_group_excel(request, group_id):
    if not request.user.is_authenticated:
        return redirect('/api/login/')
    from main.models import Group
    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        return HttpResponse("Guruh topilmadi", status=404)
    student_users = User.objects.filter(role='student', group=group)
    student_tests = StudentTest.objects.filter(completed=True, student__in=student_users).select_related('student', 'test', 'test__subject', 'test__group').order_by('student__username', '-start_time')
    wb = Workbook()
    ws = wb.active
    ws.title = f"{group.name} - Talabalar"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    headers = ["Talaba F.I.Sh.", "Username", "Test", "Fan", "Test sanasi", "Savollar soni", "To'g'ri javob", "Xato javob", "Ball", "Maksimal ball", "Foiz"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    row = 2
    for stest in student_tests:
        fio = f"{stest.student.first_name} {stest.student.last_name}"
        username = stest.student.username
        test_name = stest.test.subject.name if stest.test.subject else "-"
        subject = stest.test.subject.name if stest.test.subject else "-"
        test_date = stest.start_time.strftime("%d.%m.%Y %H:%M")
        question_ids = stest.question_ids if stest.question_ids else []
        if question_ids:
            answers = StudentAnswer.objects.filter(student_test=stest, question_id__in=question_ids)
        else:
            answers = StudentAnswer.objects.filter(student_test=stest)
        total = answers.count()
        correct = answers.filter(is_correct=True).count()
        incorrect = total - correct
        score = sum([a.score for a in answers])
        percent = int((correct / total) * 100) if total else 0
        data = [fio, username, test_name, subject, test_date, total, correct, incorrect, score, stest.test.total_score, f"{percent}%"]
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1
    column_widths = [22, 15, 20, 18, 18, 10, 12, 12, 10, 12, 8]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{group.name}_talabalar_natijalari.xlsx"'
    wb.save(response)
    return response
from django.urls import reverse
# Django templates orqali API test qilish uchun viewlar
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse, HttpResponse
from main.models import Test, Question, AnswerOption, StudentTest, StudentAnswer
from main.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone
import io

# Login sahifasi
def testapi_login(request):
    if request.method == 'POST':
        access_code = request.POST.get('access_code')
        from main.models import User
        try:
            user = User.objects.get(access_code=access_code, role='student')
            user.backend = 'django.contrib.auth.backends.ModelBackend'
            login(request, user)
            return redirect('testapi_dashboard')
        except User.DoesNotExist:
            return render(request, 'test_api/login.html', {'error': 'Access code xato yoki student topilmadi!'})
    return render(request, 'test_api/login.html')

# Dashboard sahifasi
def testapi_dashboard(request):
    if not request.user.is_authenticated:
        return redirect('/api/login/')
    
    now = timezone.now()
    if hasattr(request.user, 'group') and request.user.group:
        # Faqat muddati tugamagan testlar (end_time bo'lmasa, test ko'rinadi)
        tests = Test.objects.filter(group=request.user.group, active=True)
        filtered_tests = []
        for test in tests:
            if test.end_time:
                if now <= test.end_time:
                    filtered_tests.append(test)
            else:
                filtered_tests.append(test)
        tests = filtered_tests
    else:
        tests = []

    # Har bir fan/guruh/semestr uchun faqat eng so‘nggi (yoki eng muhim) test ko‘rsatiladi
    from collections import OrderedDict
    test_map = OrderedDict()  # (subject_id, group_id, semester_id) -> test
    for test in sorted(tests, key=lambda t: t.created_at, reverse=True):
        # Testga bog‘liq semester aniqlash
        from main.models import GroupSubject
        semester_id = None
        gs = GroupSubject.objects.filter(group=test.group, subject=test.subject).first()
        if gs:
            semester_id = gs.semester_id
        key = (test.subject_id, test.group_id, semester_id)
        if key not in test_map:
            test_map[key] = test

    final_tests = list(test_map.values())
    test_statuses = {}
    for test in final_tests:
        # Testga bog‘liq semester aniqlash
        from main.models import GroupSubject
        semester_id = None
        gs = GroupSubject.objects.filter(group=test.group, subject=test.subject).first()
        if gs:
            semester_id = gs.semester_id
        filter_kwargs = dict(
            student=request.user,
            group_id=test.group_id,
            subject_id=test.subject_id,
            completed=True,
            can_retake=False
        )
        if semester_id is not None:
            filter_kwargs['semester_id'] = semester_id
        else:
            filter_kwargs['semester_id__isnull'] = True
        participated = StudentTest.objects.filter(**filter_kwargs).exists()
        if participated:
            test_statuses[test.id] = 'done'
        else:
            test_statuses[test.id] = 'new'

    return render(request, 'test_api/dashboard.html', {'tests': final_tests, 'test_statuses': test_statuses})


# Test savollari va javob berish
def testapi_test(request, test_id):
    if not request.user.is_authenticated:
        return redirect('/api/login/')

    from main.models import TestQuestion, StudentAnswer, StudentTest
    import random
    test = Test.objects.get(id=test_id)
    test_questions = list(TestQuestion.objects.filter(test=test))
    # Talaba shu testda qatnashganmi?
    # Talaba shu testda qatnashganmi? (fan, guruh, semestr bo'yicha faqat 1 marta)
    from main.models import GroupSubject
    group = test.group
    subject = test.subject
    semester = None
    gs = GroupSubject.objects.filter(group=group, subject=subject).first()
    if gs:
        semester = gs.semester
    old_test = StudentTest.objects.filter(
        student=request.user,
        group=group,
        subject=subject,
        semester=semester,
        completed=True,
        can_retake=False
    ).first()
    if old_test:
        return render(request, 'test_api/already_participated.html', {'test': test})
    # Har bir talaba uchun random savollar tanlash (test.question_count ta)
    if not request.session.get(f'test_{test.id}_question_ids'):
        selected_tqs = random.sample(test_questions, min(test.question_count, len(test_questions)))
        question_ids = [tq.question.id for tq in selected_tqs]
        request.session[f'test_{test.id}_question_ids'] = question_ids
    else:
        question_ids = request.session[f'test_{test.id}_question_ids']
        selected_tqs = [tq for tq in test_questions if tq.question.id in question_ids]
    questions = [tq.question for tq in selected_tqs]
    answered_questions = []
    # Matching uchun: har bir matching savol uchun aralashtirilgan right variantlar
    matching_rights_dict = {}
    for q in questions:
        if q.question_type == 'matching':
            rights = [opt for opt in q.answer_options.all() if opt.right or opt.image]
            random.shuffle(rights)
            matching_rights_dict[q.id] = rights

    if request.method == 'POST':
        question_ids = request.session.get(f'test_{test.id}_question_ids', [])
        selected_tqs = [tq for tq in test_questions if tq.question.id in question_ids]
        stest = StudentTest.objects.create(
            student=request.user,
            test=test,
            group=group,
            subject=subject,
            semester=semester,
            question_ids=question_ids
        )
        for q in questions:
            is_correct = False
            score = 0
            tq = next((tq for tq in selected_tqs if tq.question.id == q.id), None)
            if q.question_type == 'single_choice':
                ans_id = request.POST.get(f'question_{q.id}')
                option = AnswerOption.objects.filter(id=ans_id).first()
                sa = StudentAnswer.objects.create(student_test=stest, question=q)
                if option:
                    sa.answer_option.add(option)
                    if option.is_correct:
                        is_correct = True
                        score = tq.score if tq else 0
                sa.is_correct = is_correct
                sa.score = score
                sa.save()
                answered_questions.append(q.id)
            elif q.question_type == 'true_false':
                ans_val = request.POST.get(f'question_{q.id}')
                correct_option = q.answer_options.filter(is_correct=True).first()
                sa = StudentAnswer.objects.create(student_test=stest, question=q, text_answer=ans_val)
                if correct_option and (
                    (ans_val == 'true' and correct_option.text.strip().lower() in ['to‘g‘ri', 'to‘gri', 'true']) or
                    (ans_val == 'false' and correct_option.text.strip().lower() in ['noto‘g‘ri', 'noto‘gri', 'false'])
                ):
                    is_correct = True
                    score = tq.score if tq else 0
                sa.is_correct = is_correct
                sa.score = score
                sa.save()
                answered_questions.append(q.id)
            elif q.question_type == 'multiple_choice':
                selected = [opt for opt in q.answer_options.all() if request.POST.get(f'question_{q.id}_{opt.id}')]
                correct_options = list(q.answer_options.filter(is_correct=True))
                sa = StudentAnswer.objects.create(student_test=stest, question=q)
                for opt in selected:
                    sa.answer_option.add(opt)
                if set(selected) == set(correct_options):
                    is_correct = True
                    score = tq.score if tq else 0
                sa.is_correct = is_correct
                sa.score = score
                sa.save()
                answered_questions.append(q.id)
            elif q.question_type == 'fill_in_blank':
                txt = request.POST.get(f'question_{q.id}', '').strip().lower()
                correct_option = q.answer_options.filter(is_correct=True).first()
                if correct_option and txt == correct_option.text.strip().lower():
                    is_correct = True
                    score = tq.score if tq else 0
                StudentAnswer.objects.create(
                    student_test=stest,
                    question=q,
                    text_answer=txt,
                    is_correct=is_correct,
                    score=score
                )
                answered_questions.append(q.id)
            elif q.question_type == 'matching':
                correct = True
                left_options = [opt for opt in q.answer_options.all() if opt.left]
                for idx, left_opt in enumerate(left_options, 1):
                    selected_right_id = request.POST.get(f'matching_{q.id}_{idx}')
                    if not selected_right_id:
                        correct = False
                    else:
                        right_opt = AnswerOption.objects.filter(id=selected_right_id).first()
                        if not right_opt:
                            correct = False
                        else:
                            # Matnli yoki rasmli javobni tekshirish
                            if left_opt.right and right_opt.right:
                                if left_opt.right != right_opt.right:
                                    correct = False
                            elif left_opt.image and right_opt.image:
                                if left_opt.image.name != right_opt.image.name:
                                    correct = False
                            else:
                                correct = False
                if correct:
                    is_correct = True
                    score = tq.score if tq else 0
                StudentAnswer.objects.create(
                    student_test=stest,
                    question=q,
                    is_correct=is_correct,
                    score=score
                )
                answered_questions.append(q.id)
            elif q.question_type == 'sentence_ordering':
                txt = request.POST.get(f'question_{q.id}', '').strip()
                correct_options = q.answer_options.filter(is_correct=True).order_by('id')
                student_words = [w.strip().lower() for w in txt.split()]
                correct_words = [opt.text.strip().lower() for opt in correct_options]
                if student_words == correct_words:
                    is_correct = True
                    score = tq.score if tq else 0
                StudentAnswer.objects.create(
                    student_test=stest,
                    question=q,
                    text_answer=txt,
                    is_correct=is_correct,
                    score=score
                )
                answered_questions.append(q.id)
        
        # Testni tugallangan deb belgilash
        stest.completed = True
        stest.total_score = sum([a.score for a in StudentAnswer.objects.filter(student_test=stest)])
        stest.save()
        
        return redirect('testapi_result', stest.id)

    else:
        stest = StudentTest.objects.filter(student=request.user, test=test).order_by('-start_time').first()
        if stest:
            answered_questions = list(stest.answers.values_list('question_id', flat=True))
    return render(request, 'test_api/test.html', {
        'test': test,
        'questions': questions,
        'answered_questions': answered_questions,
        'matching_rights_dict': matching_rights_dict,
        'test_minutes': test.minutes
    })



# Natija sahifasi
def testapi_result(request, stest_id):
    if not request.user.is_authenticated:
        return redirect('/api/login/')
    stest = StudentTest.objects.get(id=stest_id)
    
    # StudentTest modelidagi question_ids dan foydalanish
    question_ids = stest.question_ids if stest.question_ids else []
    if not question_ids:
        # Fallback: sessiondan olish
        question_ids = request.session.get(f'test_{stest.test.id}_question_ids', [])
    
    if question_ids:
        answers = StudentAnswer.objects.filter(student_test=stest, question_id__in=question_ids)
    else:
        answers = StudentAnswer.objects.filter(student_test=stest)
    
    total = answers.count()
    correct = answers.filter(is_correct=True).count()
    incorrect = total - correct
    score = sum([a.score for a in answers])
    percent = int((correct / total) * 100) if total else 0
    
    # Har bir javob uchun to'liq ma'lumot tayyorlash
    detailed_answers = []
    for answer in answers:
        question = answer.question
        user_answer = ""
        correct_answer = ""
        
        if question.question_type == 'single_choice':
            if answer.answer_option.exists():
                user_answer = answer.answer_option.first().text
            correct_option = question.answer_options.filter(is_correct=True).first()
            correct_answer = correct_option.text if correct_option else ""
            
        elif question.question_type == 'multiple_choice':
            user_answer = ", ".join([opt.text for opt in answer.answer_option.all()])
            correct_answer = ", ".join([opt.text for opt in question.answer_options.filter(is_correct=True)])
            
        elif question.question_type in ['fill_in_blank', 'true_false', 'sentence_ordering']:
            user_answer = answer.text_answer or ""
            correct_option = question.answer_options.filter(is_correct=True).first()
            correct_answer = correct_option.text if correct_option else ""
            
        elif question.question_type == 'matching':
            user_answer = "Moslashtirish javobi"
            correct_answer = "To'g'ri moslashtirish"
        
        detailed_answers.append({
            'question': question,
            'user_answer': user_answer,
            'correct_answer': correct_answer,
            'is_correct': answer.is_correct,
            'score': answer.score
        })
    
    return render(request, 'test_api/result.html', {
        'stest': stest,
        'total': total,
        'correct': correct,
        'incorrect': incorrect,
        'score': score,
        'percent': percent,
        'answers': answers,
        'detailed_answers': detailed_answers
    })

# Statistik tahlil sahifasi
def testapi_stats(request):
    if not request.user.is_authenticated:
        return redirect('/api/login/')
    stats = {
        'total_tests': Test.objects.count(),
        'total_questions': Question.objects.count(),
        'total_students': User.objects.count(),
        'total_answers': StudentAnswer.objects.count(),
    }
    return render(request, 'test_api/stats.html', {'stats': stats})

# Barcha test natijalarini ko'rish sahifasi (Admin/Controller uchun)
def testapi_all_results(request):
    if not request.user.is_authenticated:
        return redirect('/api/login/')
    # Faqat admin, controller ko'radi; override detallarini faqat superuser
    if request.user.role not in ['admin', 'controller']:
        return redirect('testapi_dashboard')
    show_override = request.user.is_superuser
    
    # Barcha tugatilgan testlarni fan va guruh bo'yicha gruppalash
    student_tests = StudentTest.objects.filter(completed=True).select_related(
        'student', 'test', 'test__subject', 'test__group'
    ).order_by('test__subject__name', 'test__group__name', 'student__username', '-start_time')
    
    # Ma'lumotlarni ierarxik tuzish: Fan -> Guruh -> Talaba -> Testlar
    organized_data = {}
    
    for stest in student_tests:
        subject_name = stest.test.subject.name if stest.test.subject else "NOMA'LUM FAN"
        group_name = stest.test.group.name if stest.test.group else "NOMA'LUM GURUH"
        group_id = stest.test.group.id if stest.test.group else 0
        student_username = stest.student.username

        # Fan bo'yicha guruhlashtirish
        if subject_name not in organized_data:
            organized_data[subject_name] = {}

        # Guruh bo'yicha guruhlashtirish
        if group_name not in organized_data[subject_name]:
            organized_data[subject_name][group_name] = {
                'group_id': group_id,
                'students': {}
            }

        # Talaba bo'yicha guruhlashtirish
        if student_username not in organized_data[subject_name][group_name]['students']:
            organized_data[subject_name][group_name]['students'][student_username] = {
                'student': stest.student,
                'tests': []
            }
        
        # Test natijalarini hisoblash
        question_ids = stest.question_ids if stest.question_ids else []
        if question_ids:
            answers = StudentAnswer.objects.filter(student_test=stest, question_id__in=question_ids)
        else:
            answers = StudentAnswer.objects.filter(student_test=stest)
        
        total = answers.count()
        correct = answers.filter(is_correct=True).count()
        incorrect = total - correct
        score = sum([a.score for a in answers])
        percent = int((correct / total) * 100) if total else 0
        
        # Har bir javob uchun batafsil ma'lumot
        answer_details = []
        for answer in answers:
            question = answer.question
            user_answer = ""
            correct_answer = ""
            
            if question.question_type == 'single_choice':
                if answer.answer_option.exists():
                    user_answer = answer.answer_option.first().text
                correct_option = question.answer_options.filter(is_correct=True).first()
                correct_answer = correct_option.text if correct_option else ""
                
            elif question.question_type == 'multiple_choice':
                user_answer = ", ".join([opt.text for opt in answer.answer_option.all()])
                correct_answer = ", ".join([opt.text for opt in question.answer_options.filter(is_correct=True)])
                
            elif question.question_type in ['fill_in_blank', 'true_false', 'sentence_ordering']:
                user_answer = answer.text_answer or ""
                correct_option = question.answer_options.filter(is_correct=True).first()
                correct_answer = correct_option.text if correct_option else ""
                
            elif question.question_type == 'matching':
                user_answer = "Moslashtirish javobi"
                correct_answer = "To'g'ri moslashtirish"
            
            answer_details.append({
                'id': answer.id,
                'question': question,
                'user_answer': user_answer,
                'correct_answer': correct_answer,
                'is_correct': answer.is_correct,
                'score': answer.score
            })
        
        # Test ma'lumotlarini qo'shish
        test_entry = {
            'student_test': stest,
            'total': total,
            'correct': correct,
            'incorrect': incorrect,
            'score': score,
            'percent': percent,
            'answer_details': answer_details,
        }
        if show_override:
            # Qo'shimcha override ma'lumotlari
            try:
                final_percent_calc = (stest.final_score / stest.test.total_score) * 100 if stest.test.total_score else percent
            except Exception:
                final_percent_calc = percent
            test_entry.update({
                'final_score': stest.final_score if hasattr(stest, 'final_score') else score,
                'is_overridden': stest.is_overridden if hasattr(stest, 'is_overridden') else False,
                'overridden_score': stest.overridden_score,
                'pass_override': stest.pass_override,
                'override_reason': stest.override_reason,
                'overridden_by': stest.overridden_by,
                'overridden_at': stest.overridden_at,
                'final_percent': int(final_percent_calc),
                'final_passed': stest.final_passed if hasattr(stest, 'final_passed') else None,
            })
        organized_data[subject_name][group_name]['students'][student_username]['tests'].append(test_entry)
    

    # Faqat completed testga ega guruhlar
    group_ids_with_results = StudentTest.objects.filter(completed=True).values_list('test__group_id', flat=True).distinct()
    groups_list = Group.objects.filter(id__in=group_ids_with_results).order_by('name')
    kafedralar_list = Kafedra.objects.all().order_by('name')
    bulimlar_list = Bulim.objects.all().order_by('name')
    return render(request, 'test_api/all_results.html', {
        'organized_data': organized_data,
        'groups_list': groups_list,
        'kafedralar_list': kafedralar_list,
        'bulimlar_list': bulimlar_list,
        'show_override': show_override
    })

# Universal Excel eksport (rolga mos)
def export_all_results_excel(request):
    if not request.user.is_authenticated:
        return redirect('/api/login/')

    user = request.user
    role = getattr(user, 'role', None)

    # Foydalanuvchi roli bo‘yicha filtr
    if role == 'admin' or role == 'controller':
        student_tests = StudentTest.objects.filter(completed=True).select_related('student', 'test', 'test__subject', 'test__group').order_by('test__subject__name', 'test__group__name', 'student__username', '-start_time')
    elif role == 'employee':
        # Xodim o‘z guruhidagi natijalarni ko‘radi
        if hasattr(user, 'group') and user.group:
            student_tests = StudentTest.objects.filter(completed=True, test__group=user.group).select_related('student', 'test', 'test__subject', 'test__group').order_by('test__subject__name', 'test__group__name', 'student__username', '-start_time')
        else:
            student_tests = StudentTest.objects.none()
    elif role == 'tutor':
        # Tutor o‘z fanidagi natijalarni ko‘radi
        if hasattr(user, 'subject') and user.subject:
            student_tests = StudentTest.objects.filter(completed=True, test__subject=user.subject).select_related('student', 'test', 'test__subject', 'test__group').order_by('test__subject__name', 'test__group__name', 'student__username', '-start_time')
        else:
            student_tests = StudentTest.objects.none()
    else:
        return redirect('testapi_dashboard')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Barcha Natijalar'

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Fan", "Guruh", "Talaba F.I.Sh.", "Username", "Test sanasi",
        "Savollar soni", "To'g'ri javob", "Xato javob", "Ball", "Maksimal ball", "Foiz"
    ]
    if request.user.is_superuser:
        headers.extend(["Yakuniy ball", "Final o'tdi?"])
    headers.extend(["Savol", "Talaba javobi", "To'g'ri javob", "Holat", "Ball (savol)"])
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    row = 2
    for stest in student_tests:
        subject = stest.test.subject.name if stest.test.subject else "NOMA'LUM FAN"
        group = stest.test.group.name if stest.test.group else "NOMA'LUM GURUH"
        student_fio = f"{stest.student.first_name} {stest.student.last_name}"
        username = stest.student.username
        test_date = stest.start_time.strftime("%d.%m.%Y %H:%M")

        question_ids = stest.question_ids if stest.question_ids else []
        if question_ids:
            answers = StudentAnswer.objects.filter(student_test=stest, question_id__in=question_ids)
        else:
            answers = StudentAnswer.objects.filter(student_test=stest)

        total = answers.count()
        correct = answers.filter(is_correct=True).count()
        incorrect = total - correct
        score = sum([a.score for a in answers])
        percent = int((correct / total) * 100) if total else 0
        final_score = getattr(stest, 'final_score', score)
        final_passed = getattr(stest, 'final_passed', None)

        if answers.exists():
            for answer in answers:
                question = answer.question
                user_answer = ""
                correct_answer = ""
                if question.question_type == 'single_choice':
                    if answer.answer_option.exists():
                        user_answer = answer.answer_option.first().text
                    correct_option = question.answer_options.filter(is_correct=True).first()
                    correct_answer = correct_option.text if correct_option else ""
                elif question.question_type == 'multiple_choice':
                    user_answer = ", ".join([opt.text for opt in answer.answer_option.all()])
                    correct_answer = ", ".join([opt.text for opt in question.answer_options.filter(is_correct=True)])
                elif question.question_type in ['fill_in_blank', 'true_false', 'sentence_ordering']:
                    user_answer = answer.text_answer or ""
                    correct_option = question.answer_options.filter(is_correct=True).first()
                    correct_answer = correct_option.text if correct_option else ""
                elif question.question_type == 'matching':
                    user_answer = "Moslashtirish javobi"
                    correct_answer = "To'g'ri moslashtirish"
                data = [subject, group, student_fio, username, test_date,
                        total, correct, incorrect, score, stest.test.total_score, f"{percent}%"]
                if request.user.is_superuser:
                    data.extend([final_score, "Ha" if final_passed else "Yo'q"])
                data.extend([question.text, user_answer, correct_answer,
                             "To'g'ri" if answer.is_correct else "Xato", answer.score])
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                row += 1
        else:
            data = [subject, group, student_fio, username, test_date,
                    0, 0, 0, 0, stest.test.total_score, "0%"]
            if request.user.is_superuser:
                data.extend([final_score, "Ha" if final_passed else "Yo'q"])
            data.extend(["Javob berilmagan", "", "", "", 0])
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    column_widths = [20, 18, 22, 15, 18, 10, 12, 12, 10, 12, 8]
    if request.user.is_superuser:
        column_widths.extend([12, 10])
    column_widths.extend([40, 30, 30, 10, 8])
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="barcha_natijalar.xlsx"'
    wb.save(response)
    return response

# Excel export funksiyasi
def export_group_results_excel(request, group_id):
    if not request.user.is_authenticated:
        return redirect('/api/login/')
    
    # Ruxsatlarni tekshirish
    from main.models import Group
    user = request.user
    role = getattr(user, 'role', None)
    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        return HttpResponse("Guruh topilmadi", status=404)

    # Admin va controller istalgan guruhni ko‘ra oladi
    if role in ['admin', 'controller']:
        student_tests = StudentTest.objects.filter(
            completed=True, test__group=group
        ).select_related('student', 'test', 'test__subject').order_by('test__subject__name', 'student__username', '-start_time')
    # Xodim faqat o‘z guruhini ko‘ra oladi
    elif role == 'employee':
        if hasattr(user, 'group') and user.group and user.group.id == group.id:
            student_tests = StudentTest.objects.filter(
                completed=True, test__group=group
            ).select_related('student', 'test', 'test__subject').order_by('test__subject__name', 'student__username', '-start_time')
        else:
            return HttpResponse("Siz faqat o‘z guruhingizni eksport qila olasiz", status=403)
    # Tutor faqat o‘z faniga tegishli guruhlarni ko‘ra oladi
    elif role == 'tutor':
        if hasattr(user, 'subject') and user.subject:
            # Guruhda shu tutor fani bormi?
            has_subject = StudentTest.objects.filter(completed=True, test__group=group, test__subject=user.subject).exists()
            if has_subject:
                student_tests = StudentTest.objects.filter(
                    completed=True, test__group=group, test__subject=user.subject
                ).select_related('student', 'test', 'test__subject').order_by('test__subject__name', 'student__username', '-start_time')
            else:
                return HttpResponse("Bu guruhda sizga tegishli fan natijalari yo‘q", status=403)
        else:
            return HttpResponse("Sizga fan biriktirilmagan", status=403)
    else:
        return redirect('testapi_dashboard')
    
    # Excel fayl yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = f"{group.name} - Test Natijalari"
    
    # Sarlavha stilini yaratish
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Sarlavhalar
    headers = [
        "Talaba Ismi", "Username", "Fan", "Test Sanasi", 
        "Jami Savollar", "To'g'ri Javoblar", "Xato Javoblar", 
        "Olingan Ball", "Maksimal Ball", "Foiz", "Savol", 
        "Talaba Javobi", "To'g'ri Javob", "Javob Holati", "Ball"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Ma'lumotlarni yozish
    row = 2
    for stest in student_tests:
        # Test natijalarini hisoblash
        question_ids = stest.question_ids if stest.question_ids else []
        if question_ids:
            answers = StudentAnswer.objects.filter(student_test=stest, question_id__in=question_ids)
        else:
            answers = StudentAnswer.objects.filter(student_test=stest)
        
        total = answers.count()
        correct = answers.filter(is_correct=True).count()
        incorrect = total - correct
        score = sum([a.score for a in answers])
        percent = int((correct / total) * 100) if total else 0
        
        # Har bir javob uchun alohida qator
        if answers.exists():
            for answer in answers:
                question = answer.question
                user_answer = ""
                correct_answer = ""
                
                if question.question_type == 'single_choice':
                    if answer.answer_option.exists():
                        user_answer = answer.answer_option.first().text
                    correct_option = question.answer_options.filter(is_correct=True).first()
                    correct_answer = correct_option.text if correct_option else ""
                    
                elif question.question_type == 'multiple_choice':
                    user_answer = ", ".join([opt.text for opt in answer.answer_option.all()])
                    correct_answer = ", ".join([opt.text for opt in question.answer_options.filter(is_correct=True)])
                    
                elif question.question_type in ['fill_in_blank', 'true_false', 'sentence_ordering']:
                    user_answer = answer.text_answer or ""
                    correct_option = question.answer_options.filter(is_correct=True).first()
                    correct_answer = correct_option.text if correct_option else ""
                    
                elif question.question_type == 'matching':
                    user_answer = "Moslashtirish javobi"
                    correct_answer = "To'g'ri moslashtirish"
                
                # Qatorga ma'lumot yozish
                data = [
                    f"{stest.student.first_name} {stest.student.last_name}",
                    stest.student.username,
                    stest.test.subject.name,
                    stest.start_time.strftime("%d.%m.%Y %H:%M"),
                    total,
                    correct,
                    incorrect,
                    score,
                    stest.test.total_score,
                    f"{percent}%",
                    question.text,
                    user_answer,
                    correct_answer,
                    "To'g'ri" if answer.is_correct else "Xato",
                    answer.score
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
        else:
            # Agar javob yo'q bo'lsa
            data = [
                f"{stest.student.first_name} {stest.student.last_name}",
                stest.student.username,
                stest.test.subject.name,
                stest.start_time.strftime("%d.%m.%Y %H:%M"),
                0, 0, 0, 0, stest.test.total_score, "0%",
                "Javob berilmagan", "", "", "", 0
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
            
            row += 1
    
    # Ustunlar kengligini sozlash
    column_widths = [20, 15, 25, 18, 12, 12, 12, 12, 12, 8, 40, 30, 30, 12, 8]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # HTTP response yaratish
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{group.name}_test_natijalari.xlsx"'
    
    # Excel faylni response ga yozish
    wb.save(response)
    return response
